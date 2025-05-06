# streamlit_app.py

import streamlit as st
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
import tempfile
import os

st.title("BOM Tracker to IP Address list Assistant")
st.markdown("### 📋 Instructions")
st.markdown("""
1. **First**, select Networked AV devices in BOM and upload the **Source Excel file** (e.g., BOM Tracker).
2. **Then**, upload the **Destination Excel file** (e.g., IP Address List).
3. Click **Run Processing** to apply the transformation.
4. Click **Download** to save the updated IP Address List file.
""")

# File upload
source_file = st.file_uploader("Upload Source Excel File (.xlsm)", type="xlsm")
destination_file = st.file_uploader("Upload Destination Excel File (.xlsm)", type="xlsm")

if st.button("Run Processing") and source_file and destination_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as tmp_src, \
         tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as tmp_dst:

        # Save uploaded files temporarily
        tmp_src.write(source_file.read())
        tmp_dst.write(destination_file.read())
        tmp_src_path = tmp_src.name
        tmp_dst_path = tmp_dst.name

    # Step 1: Load workbook and find header
    wb = load_workbook(tmp_src_path, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header_row_index = next(i for i, row in enumerate(rows) if row[0] == 'STATUS')
    headers = [str(h).replace('\n', ' ').strip() if h else '' for h in rows[header_row_index]]
    data_rows = rows[header_row_index + 1:]

    # Step 2: Clean DataFrame
    df1 = pd.DataFrame(data_rows, columns=headers)
    df1.columns = df1.columns.str.strip().str.replace(r'\s+', ' ', regex=True)

    # Step 3: Filter and rename
    filtered_df = df1[df1['AV NETWORK DEVICE'] == 'YES']
    mapped_df = filtered_df.rename(columns={
        'ROOM/AREA': 'Room #',
        'BID ITEM': 'Room Type Per AV Scope',
        'MODEL NO.': 'Device',
        'DEVICE ID': 'Device ID',
        'SERIAL NUMBER': 'Serial Number'
    })[
        ['Room Type Per AV Scope', 'Room #', 'Device', 'Device ID', 'Serial Number']
    ]

    # Step 4: Write to destination
    app = xw.App(visible=False)
    book = xw.Book(tmp_dst_path)
    sheet = book.sheets['IDF Totals']
    sheet.range((18, 1)).value = [mapped_df.columns.tolist()] + mapped_df.values.tolist()
    book.save(tmp_dst_path)
    book.close()
    app.quit()

    original_filename = destination_file.name
    # Provide download link with same filename as uploaded destination file
    with open(tmp_dst_path, "rb") as f:
        st.download_button("Download Updated Excel File", f, file_name=original_filename)

    st.success("✅ Task completed successfully!")

